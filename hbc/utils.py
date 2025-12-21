from collections import namedtuple
import datetime
from datetime import timezone
import json
import logging
from logging.handlers import RotatingFileHandler
import math
import os
from pathlib import Path
import re
import sys
import tempfile
from typing import Any, Iterable, List, Optional, Tuple, Union

from IPython import display
import pandas as pd
import yaml


def _is_console_handler(h: logging.Handler) -> bool:
    """Detect console-like handlers, including third-party ones."""
    if isinstance(h, logging.StreamHandler) and not isinstance(
        h, RotatingFileHandler
    ):
        return True
    # Some handlers aren't StreamHandler but still write to stdio.
    stream = getattr(h, "stream", None)
    return stream is sys.stdout or stream is sys.stderr


def _remove_handlers(root: logging.Logger, pred) -> None:
    for h in list(root.handlers):
        try:
            if pred(h):
                root.removeHandler(h)
                try:
                    h.close()
                except Exception:
                    pass
        except Exception:
            # why: never let handler cleanup break logging config
            continue


def clear_log(file_path: str) -> None:
    if os.path.exists(file_path):
        with open(file_path, "r+", encoding="utf-8") as f:
            f.truncate(
                0
            )  # why: ensure contents are cleared without creating new file


def get_id() -> str:
    """
    Return a unique string derived from current system time.
    """
    return datetime.datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")


def conf_log(
    file_path: str = None,
    *,
    file: bool = True,
    console: bool = True,
    level: int = logging.INFO,
    max_bytes: int = 5 * 1024 * 1024,
    backup_count: int = 5,
    reset_handlers: bool = False,
    # (filename_width, lineno_width, level_width, logger_name_width)
    field_widths: Tuple[int, int, int, int] = (20, 4, 5, 8),
) -> logging.Logger:
    """
    Configure logging with optional console and/or size-rotating file handler.
    Format: "%(asctime)s %(levelname)s %(name)s: %(message)s"
    """
    logging.disable(logging.NOTSET)  # undo any prior global disable
    if not file_path:
        file_path = get_dir_logging() / "hbc_job_generic.txt"
    clear_log(file_path)
    if file:
        print(f"Log file: {file_path}")

    root = logging.getLogger()
    root.setLevel(level)

    if reset_handlers:
        # Nuke all handlers first.
        _remove_handlers(root, lambda h: True)

    fn_w, ln_w, lvl_w, name_w = field_widths
    # Key: %-Ns for left-align strings; %Nd for right-align ints
    fmt = (
        f"%(asctime)s "
        f"%(filename)-{fn_w}s "
        f"%(lineno){ln_w}d "
        f"%(levelname)-{lvl_w}s "
        f"%(name)-{name_w}s: %(message)s"
    )

    formatter = logging.Formatter(fmt, "%Y-%m-%d %H:%M:%S")

    # --- FILE HANDLER ---
    existing_files: Iterable[logging.Handler] = [
        h for h in root.handlers if isinstance(h, RotatingFileHandler)
    ]
    if file:
        if not existing_files:
            fh = RotatingFileHandler(
                filename=file_path,
                maxBytes=max_bytes,
                backupCount=backup_count,
                encoding="utf-8",
                delay=True,
            )
            fh.setLevel(logging.NOTSET)  # root level filters
            fh.setFormatter(formatter)
            root.addHandler(fh)
        else:
            # Reconfigure all existing rotating file handlers to be consistent.
            for fh in existing_files:
                fh.setLevel(logging.NOTSET)
                fh.setFormatter(formatter)
    else:
        # Remove any file handlers when file=False.
        _remove_handlers(root, lambda h: isinstance(h, RotatingFileHandler))

    # --- CONSOLE HANDLER ---
    if console:
        # Ensure exactly one console handler.
        existing_consoles = [h for h in root.handlers if _is_console_handler(h)]
        if not existing_consoles:
            sh = logging.StreamHandler()
            sh.setLevel(logging.NOTSET)
            sh.setFormatter(formatter)
            root.addHandler(sh)
        else:
            # Keep the first, normalize it; remove extras to avoid duplicates.
            keep = existing_consoles[0]
            keep.setLevel(logging.NOTSET)
            keep.setFormatter(formatter)
            for h in existing_consoles[1:]:
                root.removeHandler(h)
                try:
                    h.close()
                except Exception:
                    pass
    else:
        # Remove ALL console-like handlers to truly silence stdout/stderr.
        _remove_handlers(root, _is_console_handler)

    # Final normalization: let root level control filtering.
    for h in root.handlers:
        if h.level > logging.NOTSET:
            h.setLevel(logging.NOTSET)

    return root


def get_config(config_name: str) -> dict[str, Any] | list[dict[str, Any]]:
    base = Path(__file__).resolve().parent / "ltp" / "configs"
    path = (
        (base / config_name)
        if Path(config_name).suffix
        else (base / f"{config_name}.yaml")
    )
    if not path.exists() and not Path(config_name).suffix:  # try .yml fallback
        path = base / f"{config_name}.yml"
    if not path.exists():
        raise FileNotFoundError(f"Config not found: {path}")
    with path.open("r", encoding="utf-8") as f:
        docs = [d for d in yaml.safe_load_all(f) if d is not None]
    if not docs:
        raise ValueError(f"Config is empty: {path}")
    return docs[0] if len(docs) == 1 else docs


def get_dir_base() -> Path:
    base = Path(tempfile.gettempdir()) / "hbc_nyc_dp"
    base.mkdir(parents=True, exist_ok=True)
    return base


def get_dir_cache(postfix: str | None = None) -> Path:
    base = get_dir_base() / "CACHE"
    cache = base / postfix if postfix else base
    cache.mkdir(parents=True, exist_ok=True)
    return cache


def get_dir_analytics() -> Path:
    analytics = get_dir_base() / "ANALYTICS"
    analytics.mkdir(parents=True, exist_ok=True)
    return analytics


def get_dir_logging() -> Path:
    logdir = get_dir_base() / "LOGS"
    logdir.mkdir(parents=True, exist_ok=True)
    return logdir


def mk_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def date_as_str(dt: datetime.date, format="%Y%m%d") -> str:
    if isinstance(dt, str):
        return dt
    return dt.strftime(format)


def date_as_iso_format(dt: datetime.date) -> str:
    return f"{dt.isoformat()}T00:00:00"


def str_as_date(dt: str):
    if dt is None:
        return
    if isinstance(dt, datetime.date):
        return dt
    return pd.to_datetime(dt).date()


def _parse_dt(s: pd.Series) -> pd.Series:
    return pd.to_datetime(s, errors="coerce")


def _nz(series: pd.Series) -> pd.Series:
    # why: treat empty/whitespace as missing for text fields
    return series.fillna("").astype(str).str.strip()


def _jsonify_unhashable(x: Any) -> Any:
    # why: make dict/list/tuple/set hashable for duplicate checks
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return None
    if isinstance(x, (dict, list, tuple, set)):
        try:
            return json.dumps(x, sort_keys=True, ensure_ascii=False)
        except Exception:
            return str(x)
    return x


def _to_hashable_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    obj_cols = [c for c in out.columns if out[c].dtype == "O"]
    for c in obj_cols:
        out[c] = out[c].map(_jsonify_unhashable)
    return out


def display_full_df(df):
    if isinstance(df, pd.Series):
        df = df.to_frame()

    pd.set_option("display.max_rows", pd.DataFrame(df).shape[0])
    pd.set_option("display.max_columns", pd.DataFrame(df).shape[1])
    pd.set_option("display.max_colwidth", 1000)
    display.display(pd.DataFrame(df))

    pd.reset_option("display.max_rows")
    pd.reset_option("display.max_columns")
    pd.reset_option("display.max_colwidth")


def to_namedtuple(d, recursive=True):
    if isinstance(d, dict):
        d = d.copy()
        if recursive:
            for k, v in d.items():
                d[k] = to_namedtuple(v, recursive)
        d = namedtuple("_", d.keys())(**d)

    return d


def cols_as_named_tuple(df: pd.DataFrame):
    dct_col = dict(zip(df.columns, df.columns))
    return to_namedtuple(dct_col)


def pretty_columns_names(df):
    df.columns = [
        val.split("[")[0].rstrip() if "[" in val else val
        for val in df.columns.values.tolist()
    ]
    df.columns = [val.lower() for val in df.columns.tolist()]
    df.columns = [val.replace("\n", "") for val in df.columns.tolist()]
    df.columns = [val.replace(" ", "_") for val in df.columns.tolist()]
    df.columns = [val.replace("(", "") for val in df.columns.tolist()]
    df.columns = [val.replace(")", "") for val in df.columns.tolist()]
    df.columns = [val.replace("&", "") for val in df.columns.tolist()]


def _sheetify(name):
    # Excel forbids : \ / ? * [ ] and max 31 chars
    clean = re.sub(r"[:\\/?*\[\]]", "_", str(name)).strip() or "Sheet1"
    return clean[:31]


def _autofit_columns(xlsx_path, sheet_name, max_width=80):
    # Set column widths to max text length in the column (capped)
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]
    for col_cells in ws.columns:
        # header-aware length
        lengths = []
        for cell in col_cells:
            v = "" if cell.value is None else str(cell.value)
            lengths.append(len(v))
        width = min(max(lengths, default=0) + 2, max_width)
        ws.column_dimensions[col_cells[0].column_letter].width = width
    wb.save(xlsx_path)
    wb.close()


def save_dataframe_as_sheet(
    dir_path,
    filename,
    df,
    sheet_name,
    replace=False,
    index=True,
    autofit=True,
    max_width=80,
):
    """
    Append DataFrame as a sheet to an Excel file, creating the file if needed.
    - replace=True overwrites an existing sheet of the same name.
    - autofit=True adjusts column widths after writing.
    """
    out_dir = Path(dir_path)
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx = out_dir / filename

    sheet = _sheetify(sheet_name)
    mode = "a" if xlsx.exists() else "w"

    if mode == "a":
        from openpyxl import load_workbook

        wb = load_workbook(xlsx)
        existing = set(wb.sheetnames)
        wb.close()

        if replace and sheet in existing:
            with pd.ExcelWriter(
                xlsx, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as w:
                df.to_excel(w, sheet_name=sheet, index=index)
            if autofit:
                _autofit_columns(xlsx, sheet, max_width=max_width)
            return

        base = sheet
        i = 1
        while sheet in existing:
            suffix = f"_{i}"
            sheet = _sheetify(base[: 31 - len(suffix)] + suffix)
            i += 1

    with pd.ExcelWriter(xlsx, engine="openpyxl", mode=mode) as w:
        df.to_excel(w, sheet_name=sheet, index=index)

    if autofit:
        _autofit_columns(xlsx, sheet, max_width=max_width)


PathLikeStr = Union[str, os.PathLike]


def path_to_str(p: Optional[PathLikeStr]) -> Optional[str]:
    """
    Convert a single Path/PathLike/str to a str.

    Parameters
    ----------
    p : Optional[PathLikeStr]
        A pathlib.Path, any os.PathLike, plain str, or None.

    Returns
    -------
    Optional[str]
        String path, or None if input was None.

    Raises
    ------
    TypeError
        If `p` is not str or PathLike.
    """
    if p is None:
        return None
    # Why: supports any os.PathLike (not only pathlib.Path).
    return os.fspath(p)


def paths_to_str(seq: Iterable[PathLikeStr]) -> List[str]:
    """
    Convert an iterable of Path/PathLike/str to a list[str].

    Notes
    -----
    - Fails fast if any element isn't a valid PathLike or str.
    """
    return [os.fspath(x) for x in seq]
